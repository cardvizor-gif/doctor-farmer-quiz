import ast
import re
from datetime import date
from pathlib import Path

from openpyxl import Workbook
from openpyxl.comments import Comment
from openpyxl.formatting.rule import FormulaRule
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.workbook.defined_name import DefinedName

PROJECT = Path('/home/ubuntu/doctor-farmer-quiz')
CATALOG_PATH = PROJECT / 'client/src/data/priceCatalog.ts'
OUTPUT = Path('/home/ubuntu/Doctor_Farmer_Регламенты_шаблон.xlsx')

# Parse the current catalog without changing it.
text = CATALOG_PATH.read_text(encoding='utf-8')
match = re.search(r"export const PRICE_CATALOG: PriceItem\[\] = (\[.*\]);\s*$", text, re.S)
if not match:
    raise RuntimeError('PRICE_CATALOG not found')
items = ast.literal_eval(match.group(1))
if len(items) != 68:
    raise RuntimeError(f'Expected 68 catalog items, found {len(items)}')

# URLs supplied by the user, mapped to exact/near-exact catalog names where unambiguous.
source_urls = {
    'Актеон, ВР': 'https://www.doctorfarmer.ru/ru/product/3/',
    'Битрин, КС': 'https://www.doctorfarmer.ru/ru/product/101/',
    'Битрин, КС* (с краской)': 'https://www.doctorfarmer.ru/ru/product/101/',
    'Богдэн, ВДГ': 'https://www.doctorfarmer.ru/ru/product/90/',
    'Вернер, КС': 'https://www.doctorfarmer.ru/ru/product/75/',
    'Гласис, ВР': 'https://www.doctorfarmer.ru/ru/product/104/',
    'Гуарил, ВДГ': 'https://www.doctorfarmer.ru/ru/product/69/',
    'Клегал, МКЭ': 'https://www.doctorfarmer.ru/ru/product/87/',
    'КлопЭфир, КЭ': 'https://www.doctorfarmer.ru/ru/product/19/',
    'КлопЭфир Интенсив': 'https://www.doctorfarmer.ru/ru/product/53/',
    'КлопЭфир Микс': 'https://www.doctorfarmer.ru/ru/product/20/',
    'Кэйталин, ВР': 'https://www.doctorfarmer.ru/ru/product/71/',
    'Кэйталин Экстра, ВДГ': 'https://www.doctorfarmer.ru/ru/product/72/',
    'Ламонд, ВР': 'https://www.doctorfarmer.ru/ru/product/76/',
    'Неон-99': 'https://www.doctorfarmer.ru/ru/product/24/',
    'Орель, ВР': 'https://www.doctorfarmer.ru/ru/product/97/',
    'ПроТэб, КМЭ': 'https://www.doctorfarmer.ru/ru/product/73/',
    'Редут, КС': 'https://www.doctorfarmer.ru/ru/product/27/',
    'Рекрут, КС': 'https://www.doctorfarmer.ru/ru/product/28/',
    'Ромул, ВДГ': 'https://www.doctorfarmer.ru/ru/product/30/',
    'Сикурс, ВР': 'https://www.doctorfarmer.ru/ru/product/48/',
    'Сотейра, ВРК': 'https://www.doctorfarmer.ru/ru/product/49/',
    'Софт, КЭ': 'https://www.doctorfarmer.ru/ru/product/77/',
    'Стратег, КС': 'https://www.doctorfarmer.ru/ru/product/32/',
    'Тайпан, КЭ': 'https://www.doctorfarmer.ru/ru/product/50/',
    'Тайсон, КС*': 'https://www.doctorfarmer.ru/ru/product/102/',
    'Тиматерр, КС (красный)': 'https://www.doctorfarmer.ru/ru/product/56/',
    'Тиматерр, КС': 'https://www.doctorfarmer.ru/ru/product/106/',
    'Триатлон Плюс': 'https://www.doctorfarmer.ru/ru/product/58/',
    'Триатлон Экстра': 'https://www.doctorfarmer.ru/ru/product/59/',
    'Турион, КЭ': 'https://www.doctorfarmer.ru/ru/product/35/',
    'Флагман, КС': 'https://www.doctorfarmer.ru/ru/product/55/',
    'Форсер Энто, КС': 'https://www.doctorfarmer.ru/ru/product/45/',
    'Фэнс, КЭ': 'https://www.doctorfarmer.ru/ru/product/79/',
    'Фэнсди, КС': 'https://www.doctorfarmer.ru/ru/product/70/',
    'Цунами, КЭ': 'https://www.doctorfarmer.ru/ru/product/38/',
    'Элант, КЭ': 'https://www.doctorfarmer.ru/ru/product/40/',
    'Элант Экстра, СЭ': 'https://www.doctorfarmer.ru/ru/product/52/',
    'Элант Премиум, КЭ': 'https://www.doctorfarmer.ru/ru/product/41/',
    'ЭтилФло, СЭ': 'https://www.doctorfarmer.ru/ru/product/98/',
}

wb = Workbook()
ws_info = wb.active
ws_info.title = 'Инструкция'
ws_ref = wb.create_sheet('Справочник')
ws_reg = wb.create_sheet('Регламенты')
ws_lists = wb.create_sheet('Списки')

# Brand-aligned light theme.
DARK = '12352A'
GREEN = '194F38'
MID = '2E7D52'
LIGHT = 'E8EFE5'
PALE = 'F4F7F1'
PALE2 = 'FBFCF9'
ORANGE = 'C77722'
RED = 'B91C1C'
GRAY = '6F7A73'
BORDER = 'D8E1D8'
INPUT = 'FFF8E7'
WHITE = 'FFFFFF'
THIN = Side(style='thin', color=BORDER)
MEDIUM = Side(style='medium', color=GREEN)

font_title = Font(name='Aptos Display', size=20, bold=True, color=DARK)
font_section = Font(name='Aptos Display', size=13, bold=True, color=GREEN)
font_header = Font(name='Aptos', size=10, bold=True, color=WHITE)
font_body = Font(name='Aptos', size=10, color=DARK)
font_note = Font(name='Aptos', size=9, italic=True, color=GRAY)

for ws in (ws_info, ws_ref, ws_reg, ws_lists):
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = 'B6'
    ws.column_dimensions['A'].width = 3

# ---------- Instructions ----------
ws_info['B2'] = 'Шаблон регламентов применения — Doctor Farmer'
ws_info['B2'].font = font_title
ws_info.merge_cells('B2:H2')
ws_info['B3'] = 'Заполняйте только жёлтые поля на листе «Регламенты». Серые поля подтягиваются из справочника автоматически.'
ws_info['B3'].font = font_note
ws_info.merge_cells('B3:H3')

ws_info['B5'] = 'Быстрый контроль'
ws_info['B5'].font = font_section
summary = [
    ('Позиций в каталоге', "=COUNTA('Справочник'!$B$6:$B$73)"),
    ('Позиций с заполненным регламентом', "=COUNTIF('Справочник'!$G$6:$G$73,\"Заполнено\")"),
    ('Позиций в работе', "=COUNTIF('Справочник'!$G$6:$G$73,\"В работе\")"),
    ('Позиций без данных', "=COUNTIF('Справочник'!$G$6:$G$73,\"Нет данных\")"),
]
for i, (label, formula) in enumerate(summary, start=6):
    ws_info.cell(i, 2, label).font = Font(name='Aptos', size=10, bold=True, color=GRAY)
    ws_info.cell(i, 3, formula).font = Font(name='Aptos', size=12, bold=True, color=GREEN)
    ws_info.cell(i, 2).fill = PatternFill('solid', fgColor=PALE)
    ws_info.cell(i, 3).fill = PatternFill('solid', fgColor=LIGHT)
    ws_info.cell(i, 2).border = ws_info.cell(i, 3).border = Border(bottom=THIN)

ws_info['B12'] = 'Как заполнять'
ws_info['B12'].font = font_section
instructions = [
    '1. На листе «Регламенты» одна строка должна описывать одну комбинацию: препарат + культура/назначение + конкретный объект или норму.',
    '2. Если у препарата разные нормы для разных культур, добавьте отдельные строки для каждой культуры. Название препарата выбирается из выпадающего списка.',
    '3. Вносите только подтверждённые сведения из официального регламента, инструкции или согласованного источника. Не объединяйте разные культуры в одну строку, если отличаются нормы или фазы.',
    '4. В поле «Фаза и время обработки» переносите формулировку из графы «Способ, время обработки, особенности применения».',
    '5. В поле «Условия и температура» указывайте только прямо подтверждённые условия. Если температура не указана, напишите: «Температурный режим не указан в источнике».',
    '6. В поле «Ограничения и требования» указывайте срок ожидания, кратность, ограничения по фазе, погоде, севообороту и совместимости — только если они есть в источнике.',
    '7. В поле «Источник / ссылка» вставьте ссылку на страницу или название документа. В поле «Статус» выберите «Заполнено» только после проверки всех обязательных полей.',
]
for i, text_line in enumerate(instructions, start=13):
    ws_info.cell(i, 2, text_line).font = font_body
    ws_info.cell(i, 2).alignment = Alignment(wrap_text=True, vertical='top')
    ws_info.merge_cells(start_row=i, start_column=2, end_row=i, end_column=8)
    ws_info.row_dimensions[i].height = 32 if i in (15, 16, 17) else 25

ws_info['B22'] = 'Навигация'
ws_info['B22'].font = font_section
for row, sheet_name in enumerate(['Справочник', 'Регламенты', 'Списки'], start=23):
    cell = ws_info.cell(row, 2, sheet_name)
    cell.hyperlink = f"#'{sheet_name}'!A1"
    cell.font = Font(name='Aptos', size=10, bold=True, color=MID, underline='single')

ws_info['B28'] = 'Цветовая легенда'
ws_info['B28'].font = font_section
legend = [('Серые поля', 'Данные каталога, не редактировать', BORDER), ('Жёлтые поля', 'Поля для заполнения', INPUT), ('Зелёный статус', 'Строка проверена и готова к переносу', 'E8F5E9'), ('Красный статус', 'Нужна проверка или источник', 'FFEBEE')]
for i, (label, desc, color) in enumerate(legend, start=29):
    ws_info.cell(i, 2, label).fill = PatternFill('solid', fgColor=color)
    ws_info.cell(i, 3, desc).font = font_body
    ws_info.merge_cells(start_row=i, start_column=3, end_row=i, end_column=6)

ws_info.column_dimensions['B'].width = 30
for col in 'CDEFGH':
    ws_info.column_dimensions[col].width = 18
ws_info.freeze_panes = None

# ---------- Reference sheet ----------
ref_headers = ['№', 'Препарат', 'Группа', 'Действующее вещество', 'Культуры / назначение', 'Норма из каталога', 'Статус регламента', 'Предоставленный источник']
for col, header in enumerate(ref_headers, start=2):
    cell = ws_ref.cell(5, col, header)
    cell.font = font_header
    cell.fill = PatternFill('solid', fgColor=GREEN)
    cell.alignment = Alignment(wrap_text=True, vertical='center')
    cell.border = Border(top=MEDIUM, bottom=MEDIUM)
ws_ref.row_dimensions[5].height = 32

for row_num, item in enumerate(items, start=6):
    values = [row_num - 5, item['name'], item['group'], item['dv'], '; '.join(item['cultures']), item['rate']]
    for col, value in enumerate(values, start=2):
        cell = ws_ref.cell(row_num, col, value)
        cell.font = font_body
        cell.alignment = Alignment(wrap_text=True, vertical='top')
        cell.fill = PatternFill('solid', fgColor='F1F4F1')
        cell.border = Border(bottom=THIN)
    ws_ref.cell(row_num, 8, f'=IF(COUNTIFS(\'Регламенты\'!$B$6:$B$205,B{row_num},\'Регламенты\'!$Q$6:$Q$205,"Заполнено")>0,"Заполнено",IF(COUNTIF(\'Регламенты\'!$B$6:$B$205,B{row_num})>0,"В работе","Нет данных"))')
    ws_ref.cell(row_num, 8).font = font_body
    ws_ref.cell(row_num, 8).alignment = Alignment(horizontal='center', vertical='center')
    ws_ref.cell(row_num, 8).fill = PatternFill('solid', fgColor=PALE2)
    ws_ref.cell(row_num, 8).border = Border(bottom=THIN)
    url_cell = ws_ref.cell(row_num, 9, source_urls.get(item['name'], ''))
    url_cell.font = Font(name='Aptos', size=9, color=MID, underline='single' if url_cell.value else None)
    url_cell.alignment = Alignment(wrap_text=True, vertical='top')
    if url_cell.value:
        url_cell.hyperlink = url_cell.value
    url_cell.border = Border(bottom=THIN)
    ws_ref.row_dimensions[row_num].height = 38

ref_widths = {'B': 6, 'C': 26, 'D': 28, 'E': 52, 'F': 32, 'G': 18, 'H': 22, 'I': 46}
for col, width in ref_widths.items():
    ws_ref.column_dimensions[col].width = width
ws_ref.freeze_panes = 'C6'
ref_table = Table(displayName='ReferenceCatalog', ref='B5:I73')
ref_table.tableStyleInfo = TableStyleInfo(name='TableStyleMedium4', showFirstColumn=False, showLastColumn=False, showRowStripes=True, showColumnStripes=False)
ws_ref.add_table(ref_table)

# ---------- Regulations input sheet ----------
reg_headers = ['№', 'Препарат', 'Действующее вещество', 'Группа', 'Культура / назначение', 'Вредный объект', 'Фаза и время обработки', 'Норма расхода препарата', 'Единица нормы', 'Расход рабочей жидкости', 'Условия применения', 'Температура / погода', 'Срок ожидания, дней', 'Кратность обработок', 'Ограничения и требования', 'Источник / ссылка', 'Статус', 'Примечание']
for col, header in enumerate(reg_headers, start=2):
    cell = ws_reg.cell(5, col, header)
    cell.font = font_header
    cell.fill = PatternFill('solid', fgColor=GREEN)
    cell.alignment = Alignment(wrap_text=True, vertical='center')
    cell.border = Border(top=MEDIUM, bottom=MEDIUM)
    cell.comment = Comment('Заполняйте только жёлтые поля. Одну строку используйте для одной комбинации культура/объект/норма.', 'Doctor Farmer')
ws_reg.row_dimensions[5].height = 46

# First 68 rows are pre-seeded with one editable line per catalog item; add 132 blank lines for additional crop-specific rules.
for row_num in range(6, 206):
    item = items[row_num - 6] if row_num - 6 < len(items) else None
    ws_reg.cell(row_num, 2, row_num - 5)
    if item:
        ws_reg.cell(row_num, 3, item['name'])
    # Formula lookups for catalog metadata.
    ws_reg.cell(row_num, 4, f'=IFERROR(VLOOKUP(C{row_num},\'Справочник\'!$C$6:$G$73,3,FALSE),"")')
    ws_reg.cell(row_num, 5, f'=IFERROR(VLOOKUP(C{row_num},\'Справочник\'!$C$6:$G$73,2,FALSE),"")')
    for col in range(2, 20):
        cell = ws_reg.cell(row_num, col)
        cell.font = font_body
        cell.alignment = Alignment(wrap_text=True, vertical='top')
        cell.border = Border(bottom=THIN)
        if col in (2, 3, 4, 5):
            cell.fill = PatternFill('solid', fgColor='F1F4F1')
        else:
            cell.fill = PatternFill('solid', fgColor=INPUT)
    ws_reg.cell(row_num, 17, 'Черновик')
    ws_reg.cell(row_num, 17).fill = PatternFill('solid', fgColor=INPUT)
    ws_reg.row_dimensions[row_num].height = 54

reg_widths = {'B': 6, 'C': 25, 'D': 38, 'E': 22, 'F': 26, 'G': 34, 'H': 38, 'I': 19, 'J': 14, 'K': 22, 'L': 34, 'M': 30, 'N': 16, 'O': 16, 'P': 38, 'Q': 42, 'R': 18, 'S': 34}
for col, width in reg_widths.items():
    ws_reg.column_dimensions[col].width = width
ws_reg.freeze_panes = 'F6'
reg_table = Table(displayName='RegulationsInput', ref='B5:S205')
reg_table.tableStyleInfo = TableStyleInfo(name='TableStyleMedium4', showFirstColumn=False, showLastColumn=False, showRowStripes=True, showColumnStripes=False)
ws_reg.add_table(reg_table)

# Input validation.
products_ref = "'Списки'!$A$2:$A$69"
wb.defined_names.add(DefinedName('ProductList', attr_text=products_ref))
product_dv = DataValidation(type='list', formula1='=ProductList', allow_blank=True)
product_dv.error = 'Выберите препарат из списка на листе «Списки».'
product_dv.errorTitle = 'Неверное название препарата'
product_dv.prompt = 'Выберите препарат из выпадающего списка.'
product_dv.promptTitle = 'Препарат'
ws_reg.add_data_validation(product_dv)
product_dv.add('C6:C205')

unit_dv = DataValidation(type='list', formula1='=UnitList', allow_blank=True)
ws_reg.add_data_validation(unit_dv)
unit_dv.add('J6:J205')
status_dv = DataValidation(type='list', formula1='=StatusList', allow_blank=False)
ws_reg.add_data_validation(status_dv)
status_dv.add('R6:R205')

# Numeric validation only where a number is expected; textual rates are allowed in the main field.
int_dv = DataValidation(type='whole', operator='between', formula1='0', formula2='365', allow_blank=True)
int_dv.error = 'Введите количество дней целым числом от 0 до 365.'
ws_reg.add_data_validation(int_dv)
int_dv.add('N6:N205')

# Status highlighting.
ws_reg.conditional_formatting.add('R6:R205', FormulaRule(formula=['$R6="Заполнено"'], fill=PatternFill('solid', fgColor='E8F5E9')))
ws_reg.conditional_formatting.add('R6:R205', FormulaRule(formula=['$R6="Нужна проверка"'], fill=PatternFill('solid', fgColor='FFEBEE')))
ws_reg.conditional_formatting.add('R6:R205', FormulaRule(formula=['$R6="Черновик"'], fill=PatternFill('solid', fgColor='FFF3E0')))

# ---------- Lists sheet ----------
ws_lists['A1'] = 'Препараты'
ws_lists['B1'] = 'Единицы нормы'
ws_lists['C1'] = 'Статусы'
for cell in ws_lists[1]:
    cell.font = font_header
    cell.fill = PatternFill('solid', fgColor=GREEN)
for i, item in enumerate(items, start=2):
    ws_lists.cell(i, 1, item['name'])
units = ['л/га', 'л/т', 'кг/га', 'кг/т', 'г/га', 'г/т', 'мл/га', 'мл/т', 'л/1000 л', 'кг/1000 л', 'другое']
for i, value in enumerate(units, start=2):
    ws_lists.cell(i, 2, value)
statuses = ['Черновик', 'Заполнено', 'Нужна проверка']
for i, value in enumerate(statuses, start=2):
    ws_lists.cell(i, 3, value)
wb.defined_names.add(DefinedName('UnitList', attr_text="'Списки'!$B$2:$B$12"))
wb.defined_names.add(DefinedName('StatusList', attr_text="'Списки'!$C$2:$C$4"))
ws_lists.sheet_state = 'hidden'

# Workbook polish.
for ws in (ws_info, ws_ref, ws_reg):
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.page_margins.left = 0.25
    ws.page_margins.right = 0.25
    ws.page_margins.top = 0.5
    ws.page_margins.bottom = 0.5
    ws.sheet_properties.outlinePr.summaryBelow = True

ws_ref.auto_filter.ref = 'B5:I73'
ws_reg.auto_filter.ref = 'B5:S205'
wb.active = 0
wb.properties.creator = 'Doctor Farmer'
wb.properties.title = 'Шаблон регламентов применения препаратов'
wb.properties.subject = 'Заполнение регламентов для Базы знаний Doctor Farmer'
wb.properties.description = 'Шаблон для подтвержденных агрономических регламентов по каталогу из 68 позиций.'
wb.calculation.fullCalcOnLoad = True
wb.calculation.forceFullCalc = True
wb.save(OUTPUT)
print(f'Created: {OUTPUT}')
print(f'Catalog rows: {len(items)}')
