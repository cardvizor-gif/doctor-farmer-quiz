import ast
import re
from pathlib import Path
from openpyxl import load_workbook

upload = Path('/home/ubuntu/upload/Doctor_Farmer_Регламенты_заполнено.xlsx')
project = Path('/home/ubuntu/doctor-farmer-quiz/client/src/data/priceCatalog.ts')
wb = load_workbook(upload, data_only=False)
ref = wb['Справочник']
file_names = [str(ref.cell(row, 3).value).strip() for row in range(6, ref.max_row + 1) if ref.cell(row, 3).value]
text = project.read_text(encoding='utf-8')
match = re.search(r"export const PRICE_CATALOG: PriceItem\[\] = (\[.*\]);\s*$", text, re.S)
items = ast.literal_eval(match.group(1))
project_names = [item['name'] for item in items]
print('FILE_REFERENCE_COUNT', len(file_names))
print('PROJECT_COUNT', len(project_names))
print('MISSING_IN_FILE')
for name in project_names:
    if name not in file_names:
        print(name)
print('UNKNOWN_IN_FILE')
for name in file_names:
    if name not in project_names:
        print(name)
print('FILE_DUPLICATE_NAMES')
from collections import Counter
for name, count in Counter(file_names).items():
    if count > 1:
        print(name, count)
