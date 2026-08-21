from pathlib import Path
from openpyxl import load_workbook

path = Path('/home/ubuntu/upload/Doctor_Farmer_Регламенты_заполнено.xlsx')
wb = load_workbook(path, data_only=False)
ws = wb['Регламенты']
headers = [ws.cell(5, c).value for c in range(1, ws.max_column + 1)]
print('HEADERS:')
for i, value in enumerate(headers, 1):
    if value:
        print(i, repr(value))
print('ROWS_WITH_ANY_INPUT:')
for row in range(6, ws.max_row + 1):
    values = [ws.cell(row, c).value for c in range(2, ws.max_column + 1)]
    if any(v not in (None, '') for v in values[5:]):
        print(row, [str(v)[:120] if v is not None else '' for v in values])
        if row > 35:
            break
print('LAST_NONEMPTY_ROWS:')
seen = 0
for row in range(ws.max_row, 5, -1):
    values = [ws.cell(row, c).value for c in range(2, ws.max_column + 1)]
    if any(v not in (None, '') for v in values):
        print(row, [str(v)[:120] if v is not None else '' for v in values])
        seen += 1
        if seen >= 10:
            break
