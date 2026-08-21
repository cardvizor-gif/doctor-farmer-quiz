import json
import re
from collections import Counter, defaultdict
from pathlib import Path

from openpyxl import load_workbook

UPLOAD = Path('/home/ubuntu/upload/Doctor_Farmer_Регламенты_заполнено.xlsx')
CATALOG = Path('/home/ubuntu/doctor-farmer-quiz/client/src/data/priceCatalog.ts')
REPORT = Path('/home/ubuntu/doctor-farmer-quiz/uploaded_regulations_audit.json')

if not UPLOAD.exists():
    raise SystemExit(f'Uploaded workbook not found: {UPLOAD}')

# Read catalog names from the generated audit JSON if available; otherwise use workbook reference sheet.
wb = load_workbook(UPLOAD, data_only=False, read_only=False)
result = {'file': str(UPLOAD), 'sheets': wb.sheetnames, 'sheet_dimensions': {ws.title: [ws.max_row, ws.max_column] for ws in wb.worksheets}}

sheet_name = 'Регламенты' if 'Регламенты' in wb.sheetnames else wb.sheetnames[0]
ws = wb[sheet_name]
headers = {ws.cell(5, c).value: c for c in range(1, ws.max_column + 1) if ws.cell(5, c).value}
required_headers = ['Препарат', 'Культура / назначение', 'Вредный объект', 'Фаза и время обработки', 'Норма расхода препарата', 'Единица нормы', 'Источник / ссылка', 'Статус']
missing_headers = [h for h in required_headers if h not in headers]
if missing_headers:
    raise SystemExit(f'Missing headers: {missing_headers}')

# Load catalog names from reference sheet when available.
catalog_names = []
if 'Справочник' in wb.sheetnames:
    ref = wb['Справочник']
    for row in range(6, ref.max_row + 1):
        value = ref.cell(row, 3).value
        if value:
            catalog_names.append(str(value).strip())

if not catalog_names:
    # Fallback to current catalog file, without modifying it.
    text = CATALOG.read_text(encoding='utf-8')
    catalog_names = re.findall(r"'name': '([^']+)'", text)

# Rows with a product name are data rows; blank rows are ignored.
rows = []
for row in range(6, ws.max_row + 1):
    product = ws.cell(row, headers['Препарат']).value
    if product is None or str(product).strip() == '':
        continue
    values = {header: ws.cell(row, col).value for header, col in headers.items()}
    rows.append({'row': row, **values})

# Required content for a row to be eligible for direct transfer. Conditions/temperature/restrictions
# are checked separately because a source may explicitly say that a field is not specified.
required_content = ['Препарат', 'Культура / назначение', 'Вредный объект', 'Фаза и время обработки', 'Норма расхода препарата', 'Единица нормы', 'Источник / ссылка']
optional_but_recommended = ['Расход рабочей жидкости', 'Условия применения', 'Температура / погода', 'Срок ожидания, дней', 'Кратность обработок', 'Ограничения и требования']
placeholder_tokens = ['Требует уточнения', 'Вредные объекты по регламенту', 'Применять по рекомендациям агрономической службы', 'Соблюдать регламент безопасности и нормы расхода']

missing_by_field = Counter()
placeholder_by_field = Counter()
status_counts = Counter(str(r.get('Статус') or '').strip() or '(пусто)' for r in rows)
products = Counter(str(r['Препарат']).strip() for r in rows)
valid_status_rows = []
row_issues = []
for r in rows:
    issues = []
    for field in required_content:
        value = r.get(field)
        if value is None or str(value).strip() == '':
            missing_by_field[field] += 1
            issues.append(f'нет поля: {field}')
    for field in required_content + optional_but_recommended:
        value = str(r.get(field) or '')
        if any(token.lower() in value.lower() for token in placeholder_tokens):
            placeholder_by_field[field] += 1
            issues.append(f'шаблонное значение: {field}')
    status = str(r.get('Статус') or '').strip()
    if status == 'Заполнено' and not issues:
        valid_status_rows.append(r)
    if issues:
        row_issues.append({'row': r['row'], 'product': r.get('Препарат'), 'status': status or '(пусто)', 'issues': issues})

provided_products = set(products)
catalog_set = set(catalog_names)
missing_products = sorted(catalog_set - provided_products)
unknown_products = sorted(provided_products - catalog_set)

# A product is considered ready only if at least one row is marked Заполнено and has no required-field issues.
ready_products = sorted({str(r['Препарат']).strip() for r in valid_status_rows})
not_ready_products = sorted(catalog_set - set(ready_products))

result.update({
    'data_sheet': sheet_name,
    'header_row': 5,
    'data_rows_with_product': len(rows),
    'unique_products_in_file': len(provided_products),
    'catalog_products_from_reference': len(catalog_set),
    'status_counts': dict(status_counts),
    'rows_without_required_data': len(row_issues),
    'missing_by_field': dict(missing_by_field),
    'placeholder_by_field': dict(placeholder_by_field),
    'missing_catalog_products': missing_products,
    'unknown_products': unknown_products,
    'ready_products': ready_products,
    'not_ready_products': not_ready_products,
    'row_issues': row_issues,
    'duplicate_product_rows': {name: count for name, count in products.items() if count > 1},
    'source_url_rows': sum(1 for r in rows if str(r.get('Источник / ссылка') or '').strip().startswith(('http://', 'https://'))),
})
REPORT.write_text(json.dumps(result, ensure_ascii=False, indent=2), encoding='utf-8')

print('SHEETS:', ', '.join(wb.sheetnames))
print('DATA_SHEET:', sheet_name)
print('DATA_ROWS:', len(rows))
print('UNIQUE_PRODUCTS:', len(provided_products))
print('CATALOG_PRODUCTS:', len(catalog_set))
print('STATUS_COUNTS:', json.dumps(dict(status_counts), ensure_ascii=False))
print('READY_PRODUCTS:', len(ready_products))
print('NOT_READY_PRODUCTS:', len(not_ready_products))
print('ROWS_WITH_ISSUES:', len(row_issues))
print('MISSING_REQUIRED_BY_FIELD:', json.dumps(dict(missing_by_field), ensure_ascii=False))
print('PLACEHOLDERS_BY_FIELD:', json.dumps(dict(placeholder_by_field), ensure_ascii=False))
print('SOURCE_URL_ROWS:', result['source_url_rows'])
print('MISSING_PRODUCTS:', json.dumps(missing_products, ensure_ascii=False))
print('UNKNOWN_PRODUCTS:', json.dumps(unknown_products, ensure_ascii=False))
print('REPORT:', REPORT)
