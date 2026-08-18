import openpyxl
import json

wb = openpyxl.load_workbook("/home/ubuntu/upload/НовыйпрайсСПКУрал.xlsx", data_only=True)
sheet = wb["Прайс"]
items = []
current_cat = ""
for r in range(4, sheet.max_row + 1):
    val0 = sheet.cell(r, 1).value
    val1 = sheet.cell(r, 2).value
    val2 = sheet.cell(r, 3).value
    val3 = sheet.cell(r, 4).value
    val5 = sheet.cell(r, 6).value
    if val1 is None and val0 and isinstance(val0, str) and not str(val0).isdigit():
        current_cat = val0.strip()
        continue
    if val1 is not None and val5 is not None:
        name = str(val1).strip()
        dv = str(val2).strip() if val2 else ""
        rate = str(val3).strip() if val3 else ""
        cultures = str(val5).strip() if val5 else ""
        
        cat_lower = current_cat.lower()
        if "удобрен" in cat_lower or "микро" in cat_lower:
            group = "Удобрение"
        elif "фунгицид" in cat_lower:
            group = "Фунгицид"
        elif "инсектицид" in cat_lower:
            group = "Инсектицид"
        elif "протравит" in cat_lower:
            if "тебуконазол" in dv.lower() or "мефеноксам" in dv.lower() or "фунгицид" in cat_lower:
                group = "Фунгицидный протравитель"
            elif "имидаклоприд" in dv.lower() or "инсектицид" in cat_lower:
                group = "Инсектицидный протравитель"
            else:
                group = "Протравитель"
        elif "десикант" in cat_lower:
            group = "Десикант"
        elif "пав" in cat_lower or "адъювант" in cat_lower:
            group = "ПАВ"
        else:
            name_lower = name.lower()
            dv_lower = dv.lower()
            if "хизалофоп" in dv_lower or "клетодим" in dv_lower or "софт" in name_lower or "клегал" in name_lower or "тайпан" in name_lower:
                group = "Гербицид (злак)"
            elif "глифосат" in dv_lower:
                group = "Гербицид сплошного действия"
            else:
                group = "Гербицид"
                
        items.append({
            "name": name,
            "dv": dv,
            "rate": rate,
            "category": current_cat,
            "group": group,
            "cultures": [c.strip() for c in cultures.split(",")]
        })

out_code = "export interface PriceItem {\n  name: string;\n  dv: string;\n  rate: string;\n  category: string;\n  group: 'Гербицид' | 'Гербицид (злак)' | 'Гербицид сплошного действия' | 'Фунгицид' | 'Инсектицид' | 'Протравитель' | 'Фунгицидный протравитель' | 'Инсектицидный протравитель' | 'Удобрение' | 'Десикант' | 'ПАВ';\n  cultures: string[];\n}\n\nexport const PRICE_CATALOG: PriceItem[] = " + json.dumps(items, ensure_ascii=False, indent=2) + ";\n"

with open("/home/ubuntu/doctor-farmer-quiz/client/src/data/priceCatalog.ts", "w") as f:
    f.write(out_code)

print(f"Successfully generated priceCatalog.ts with {len(items)} items.")
