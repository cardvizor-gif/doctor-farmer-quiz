from pathlib import Path
from openpyxl import load_workbook

path = Path('/home/ubuntu/upload/Doctor_Farmer_Регламенты_заполнено.xlsx')
wb = load_workbook(path, data_only=False)
ws = wb['Регламенты']
for row in [6, 7, 8, 19, 71, 100, 150]:
    print(f'ROW {row}')
    for col in range(2, 20):
        value = ws.cell(row, col).value
        if value not in (None, ''):
            print(f'  {ws.cell(5, col).coordinate} header={ws.cell(5, col).value!r} value={value!r}')
