from pathlib import Path
from openpyxl import load_workbook

path = Path('/home/ubuntu/upload/Doctor_Farmer_Регламенты_заполнено.xlsx')
wb = load_workbook(path, data_only=False)
ws = wb['Регламенты']
headers = {ws.cell(5, c).value: c for c in range(1, ws.max_column + 1) if ws.cell(5, c).value}
fields = ['Препарат', 'Культура / назначение', 'Вредный объект', 'Фаза и время обработки', 'Норма расхода препарата', 'Единица нормы', 'Расход рабочей жидкости', 'Условия применения', 'Температура / погода', 'Срок ожидания, дней', 'Кратность обработок', 'Ограничения и требования', 'Источник / ссылка', 'Статус', 'Примечание']
for row in range(6, ws.max_row + 1):
    product = ws.cell(row, headers['Препарат']).value
    if not product:
        continue
    substantive = []
    for field in fields[1:]:
        value = ws.cell(row, headers[field]).value
        if value not in (None, '', 'Черновик'):
            substantive.append(f'{field}={value}')
    if substantive:
        print(f'ROW {row} | {product}')
        for value in substantive:
            print('  ' + value)
