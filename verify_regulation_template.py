from pathlib import Path
from openpyxl import load_workbook

path = Path('/home/ubuntu/Doctor_Farmer_Регламенты_шаблон.xlsx')
wb = load_workbook(path, data_only=False)
expected = ['Инструкция', 'Справочник', 'Регламенты', 'Списки']
assert wb.sheetnames == expected, wb.sheetnames
ws_ref = wb['Справочник']
ws_reg = wb['Регламенты']
assert ws_ref.max_row == 73, ws_ref.max_row
assert ws_reg.max_row == 205, ws_reg.max_row
assert ws_ref['B6'].value == 1
assert ws_ref['B73'].value == 68
assert ws_ref['C6'].value == 'Контур, ВР'
assert ws_reg['C6'].value == 'Контур, ВР'
assert ws_reg['D6'].value.startswith('=IFERROR(VLOOKUP')
assert len(ws_reg.data_validations.dataValidation) >= 3
assert ws_reg.auto_filter.ref == 'B5:S205'
assert wb['Списки'].sheet_state == 'hidden'
print('Workbook structure: OK')
print('Reference rows: 68')
print('Input rows: 200')
print('Data validations:', len(ws_reg.data_validations.dataValidation))
print('Hidden helper sheet: OK')
print('Output:', path)
