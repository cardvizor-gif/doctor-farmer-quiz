from collections import defaultdict
from pathlib import Path
from openpyxl import load_workbook

path = Path('/home/ubuntu/upload/Doctor_Farmer_Регламенты_заполнено.xlsx')
wb = load_workbook(path, data_only=False)
ws = wb['Регламенты']
# In the uploaded file the user appears to have entered source data in D:I; report both raw and useful content.
by_product = defaultdict(list)
for row in range(6, ws.max_row + 1):
    product = ws.cell(row, 3).value
    if not product:
        continue
    values = []
    for col in range(4, 17):
        value = ws.cell(row, col).value
        if value not in (None, '', '()', 'Черновик') and not (isinstance(value, str) and value.startswith('=')):
            values.append((ws.cell(5, col).value, value))
    if values:
        by_product[str(product).strip()].append((row, values))
print('PRODUCTS_WITH_ANY_SUBSTANTIVE_INPUT', len(by_product))
for product, rows in by_product.items():
    print(f'{product} | rows={len(rows)} | first_row={rows[0][0]}')
    for row, values in rows[:2]:
        print('  ', row, ' || '.join(f'{h}={v}' for h, v in values))
